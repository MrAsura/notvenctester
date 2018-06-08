"""
Test suite containing functions for generating test results from TestInstances
"""

import openpyxl as xl
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.formatting.rule import ColorScaleRule
#import re
import cfg
import ast

__FILE_END = r".xlsm"
__KBS = r"kbs"
__KB = r"kb"
__TIME = r"time"
__PSNR = r"psnr"
__SCALE = r"scale"
__RES = r"results"
__QPS = r"qps"
__INAMES = r"inames"

#__res_regex = r"\sProcessed\s(\d+)\sframes\sover\s(\d+)\slayer\(s\),\s*(\d+)\sbits\sAVG\sPSNR:\s(\d+[.,]\d+)\s(\d+[.,]\d+)\s(\d+[.,]\d+)"
#__lres_regex_format = r"\s\sLayer\s{lid}:\s*(\d+)\sbits,\sAVG\sPSNR:\s(\d+[.,]\d+)\s(\d+[.,]\d+)\s(\d+[.,]\d+)"
#__time_regex = r"\sEncoding\stime:\s(\d+.\d+)\ss."

__R_HEADER = ["Sequence","Layer"]
__R_HEADER_QP = "QP {}"
__R_KBS = ["Kb","Kb/s","Time (s)"]
__R_PSNR = "PSNR"
__R_PSNR_SUB = ["Y","U","V","AVG"]

__SEQ_FORMAT = "{} @ scale {}"
__PSNR_AVG = "=(6*{y}+{u}+{v})/8"

__C_AVG = r"=AVERAGE({})"
__SEQ_AVERAGE = r"Average"

__S_SEQ_HEADER = "Sequence {} results"
__S_BIT_HEADER = r"Bit comparisons"
__S_PSNR_HEADER = r"PSNR comparisons (dB)"
__S_TIME_HEADER = r"Encoding time comparisons"
__S_BDRATE_FORMAT = "=bdrate({},{},{},{},{},{},{},{},{},{},{},{},{},{},{},{})"
__S_BIT_FORMAT = "=AVERAGE({},{},{},{})/AVERAGE({},{},{},{})"
__S_TIME_FORMAT = "=AVERAGE({},{},{},{})/AVERAGE({},{},{},{})"
__S_PSNR_FORMAT = "=AVERAGE({},{},{},{})-AVERAGE({},{},{},{})"
__S_HEADER = "Result summary matrix (bdrate, bit, PSNR, Time comparisons)"

__S2_HEADER = "Result summary list"

__SR_FORMAT = r"'{sheet}'!{cell}"

__LID_TOT = -1

__LAYER2TEST_SEP = r"_layer"
__LAYER2TEST_FORMAT = r"{test}"+ __LAYER2TEST_SEP + "{lid}"

__COMBI_SEP = "+"
__LCOMBI_SEP = "_"

"""
Generate sheet layer string
"""
def makeSheetLayer(sheet,layer):
    return __LAYER2TEST_FORMAT.format(test=sheet,lid=layer)

"""
Parse sheet layer string
"""
def parseSheetLayer(string):
    return string.split(sep=__LAYER2TEST_SEP)

def parseCombiName(string):
    return string.split(sep=__COMBI_SEP)

def makeCombiName(combi):
    return __COMBI_SEP.join(combi)

def parseLayerCombiName(string):
    return string.split(sep=__LCOMBI_SEP)

def makeLayerCombiName(combi):
    return __LCOMBI_SEP.join(combi)

def getMaxLength( string_list ):
    return max((len(x) for x in string_list))

"""
Parse kb/s from test results 

def __parseKBS(res,lres,trgt,nl):
    (frames,bits) = res.group(1,3)
    lframes = int(frames)/nl
    for lid in range(nl):
        if lres[lid]:
            lbits = lres[lid].group(1)
            trgt[lid][__KBS] = float(lbits)/float(lframes)
        else:
            trgt[lid][__KBS] = float(bits)/float(frames)
    trgt[__LID_TOT][__KBS] = float(bits)/float(frames)
"""
"""
Parse kb from test results

def __parseKB(res,lres,trgt,nl):
    bits = res.group(3)
    for lid in range(nl):
        if lres[lid]:
            lbits = lres[lid].group(1)
            trgt[lid][__KB] = float(lbits)
        else:
            trgt[lid][__KB] = float(bits)
    trgt[__LID_TOT][__KB] = float(bits)
"""
"""
Parse Time

def __parseTime(res,lres,trgt,nl):
    ttime = res.group(1)
    for lid in range(nl):
        if lres[lid]:
            ltime = ttime#lres[lid].group(6)
            trgt[lid][__TIME] = float(ltime)
        else:
            trgt[lid][__TIME] = float(ttime)
    trgt[__LID_TOT][__TIME] = float(ttime)
"""
"""
Parse psnr from test results

def __parsePSNR(res,lres,trgt,nl):
    for lid in range(nl):
        if lres[lid]:
            trgt[lid][__PSNR] = lres[lid].group(2,3,4)
        else:
            trgt[lid][__PSNR] = res.group(4,5,6)
    trgt[__LID_TOT][__PSNR] = res.group(4,5,6)
"""
"""
Parse needed values

def __parseVals(res):
    trgt = {}
    res_ex = re.search(__res_regex, str(res))
    time_ex = re.search(__time_regex, str(res))
    lres_ex = {}
    num_layers = int(res_ex.group(2))
    for lid in range(num_layers):
        lres_ex[lid] = re.search(__lres_regex_format.format(lid=lid), str(res))
        trgt[lid] = {}
    trgt[__LID_TOT] = {}
    __parseKBS(res_ex,lres_ex,trgt,num_layers)
    __parseKB(res_ex,lres_ex,trgt,num_layers)
    __parseTime(time_ex,lres_ex,trgt,num_layers)
    __parsePSNR(res_ex,lres_ex,trgt,num_layers)
    return trgt
"""

"""
Build test result dict
"""
def __resBuildFunc(results,seq,qp,lid,kbs,kb,time,psnr):
    
    if not seq in results:
        results[seq] = {}
    if not qp in results[seq]:
        results[seq][qp] = {}
    if not lid in results[seq][qp]:
        results[seq][qp][lid] = {}

    results[seq][qp][lid][__KBS] = kbs
    results[seq][qp][lid][__KB] = kb
    results[seq][qp][lid][__TIME] = time
    results[seq][qp][lid][__PSNR] = psnr

"""
Sort results qps and put them to ascending order by replacing them with order numbers
@return sorted dict and a dict containing the new to old qp key mapping
"""
def __sortQps(res):
    new_res = {}
    qp_map = {}
    for (seq,qps) in res.items():
        new_res[seq] = {}
        qp_map[seq] = {}

        sorted_qps = list(map(str, sorted([ast.literal_eval(qp) for qp in qps.keys()])))
        for (i,qp) in zip(range(len(sorted_qps)),sorted_qps):
            new_res[seq][i] = qps[qp]
            qp_map[seq][i] = qp

    return (new_res,qp_map)

"""
Parse test results
@return a dict with parsed results
"""
def __parseTestResults(tests):
    results = {}
    for test in tests:
        main_res = test.getResults(__resBuildFunc,l_tot=__LID_TOT)
        (main_res,qp_names) = __sortQps(main_res)
        results[test._test_name] = {__RES: main_res, __SCALE: str(test._input_layer_scales), __QPS: qp_names, __INAMES: [__SEQ_AVERAGE,] + test.getInputNames()}
    return results

"""
Combine psnr and kbs values
"""
def __combiValues(vals):
   
    res = {__RES:{},__SCALE:'',__QPS:{}, __INAMES:vals[0][__INAMES]}
    #Init structure
    for (seq,qps) in vals[0][__RES].items():
        res[__RES][seq] = {}
        res[__QPS][seq] = {}
        for (qp,lids) in qps.items():
            res[__RES][seq][qp] = {}
            res[__QPS][seq][qp] = ""
            for (lid,val) in lids.items():
                res[__RES][seq][qp][lid] = {}
                res[__RES][seq][qp][lid][__KBS] = 0
                res[__RES][seq][qp][lid][__KB] = 0
                res[__RES][seq][qp][lid][__TIME] = 0
                res[__RES][seq][qp][lid][__PSNR] = (0,0,0)

    numv = len(vals)

    scales = []
    for item in vals:
        scales.append(item[__SCALE])
        for (seq,qps) in item[__RES].items():
            for (qp,lids) in qps.items():
                for (lid,val) in lids.items():
                    res[__RES][seq][qp][lid][__KBS] += val[__KBS]
                    res[__RES][seq][qp][lid][__KB] += val[__KB]
                    res[__RES][seq][qp][lid][__TIME] += val[__TIME]
                    res[__RES][seq][qp][lid][__PSNR] = tuple(map(lambda x,y: float(y) + float(x)/float(numv), val[__PSNR], res[__RES][seq][qp][lid][__PSNR]))
                res[__QPS][seq][qp] = makeCombiName([res[__QPS][seq][qp],item[__QPS][seq][qp]]) if len(res[__QPS][seq][qp]) > 0 else item[__QPS][seq][qp]
    res[__SCALE] = makeCombiName(scales)
        
    return res

"""
Combine psnr and kbs values in a layered fashion
"""
def __layerCombiValues(vals):
   
    res = {__RES:{},__SCALE:'',__QPS:{}, __INAMES:vals[0][__INAMES]}
    #Init structure
    for (seq,qps) in vals[0][__RES].items():
        res[__RES][seq] = {}
        res[__QPS][seq] = {}
        for qp in qps.keys():
            res[__RES][seq][qp] = {}
            res[__QPS][seq][qp] = "()"
            for lid in range(len(vals)):
                res[__RES][seq][qp][lid] = {}
                res[__RES][seq][qp][lid][__KBS] = 0
                res[__RES][seq][qp][lid][__KB] = 0
                res[__RES][seq][qp][lid][__TIME] = 0
                res[__RES][seq][qp][lid][__PSNR] = (0,0,0)
            res[__RES][seq][qp][__LID_TOT] = {}
            res[__RES][seq][qp][__LID_TOT][__KBS] = 0
            res[__RES][seq][qp][__LID_TOT][__KB] = 0
            res[__RES][seq][qp][__LID_TOT][__TIME] = 0
            res[__RES][seq][qp][__LID_TOT][__PSNR] = (0,0,0)

    numv = len(vals)

    scales = []
    for item in vals:
        scales.append(item[__SCALE])

    for (seq,qps) in res[__RES].items():
        for (qp,lids) in qps.items():
            for (lid,val) in zip(range(numv),vals):
                lids[lid][__KBS] = val[__RES][seq][qp][__LID_TOT][__KBS]
                lids[lid][__KB] = val[__RES][seq][qp][__LID_TOT][__KB]
                lids[lid][__TIME] = val[__RES][seq][qp][__LID_TOT][__TIME]
                lids[lid][__PSNR] = val[__RES][seq][qp][__LID_TOT][__PSNR]
                lids[__LID_TOT][__KBS] += val[__RES][seq][qp][__LID_TOT][__KBS]#/numv #Take the average
                lids[__LID_TOT][__KB] += val[__RES][seq][qp][__LID_TOT][__KB]
                lids[__LID_TOT][__TIME] += val[__RES][seq][qp][__LID_TOT][__TIME]
                lids[__LID_TOT][__PSNR] = tuple(map(lambda x,y: float(y) + float(x)/float(numv), val[__RES][seq][qp][__LID_TOT][__PSNR], lids[__LID_TOT][__PSNR]))

                res[__QPS][seq][qp] = str( ast.literal_eval(res[__QPS][seq][qp]) + ast.literal_eval(val[__QPS][seq][qp]) )

    res[__SCALE] = makeLayerCombiName(scales)
        
    return res

"""
Combine test results to form new tests
@param combi: list of results to combine  (or list of lists). 
@param layer_combi: list of results to combine. layer_combi[i] will be the ith layer.
"""
def __combiTestResults(results, combi, layer_combi):
    res = results.copy()

    for set in combi:
        vals = []
        for item in set:
            vals.append(results[item])
        cname = makeCombiName(set)
        res[cname] = __combiValues(vals)

    for set in layer_combi:
        vals = []
        for item in set:
            vals.append(results[item])
        cname = makeLayerCombiName(set)
        res[cname] = __layerCombiValues(vals)

    return res
    

"""
Write summary base header
@return row and col of start of data field
"""
def __writeSummaryMatrixHeader(sheet, tests, row, col):
    d_row = row + 1
    d_col = col + 1
    #Write horizontal headers/test names
    tmp_col = col + 1
    for test in tests:
        sheet.cell(row = row, column = tmp_col).value = test
        tmp_col += 1

    #Write vertical 
    tmp_row = row + 1
    for test in tests:
        sheet.cell(row = tmp_row, column = col).value = test
        tmp_row += 1

    return d_row, d_col

"""
Write summary matrix data for bdrate
"""
def __writeBDSummaryMatrix(sheet, data, row, col):
    test_col = col - 1
    test_row = row - 1
    final_r = row+len(data.keys())
    final_c = col+len(data.keys())
    for r in range(row,row+len(data.keys())):
        for c in range(col,col+len(data.keys())):
            t2 = sheet.cell(row = r, column = test_col).value
            t1 = sheet.cell(row = test_row, column = c).value
            if t1 == t2:
                sheet.cell(row = r, column = c).value = "-"
            else:
                r1 =[__SR_FORMAT.format(sheet=parseSheetLayer(t1)[0],cell=cl) for cl in data[t1][__KBS] + data[t1][__PSNR]]
                r2 =[__SR_FORMAT.format(sheet=parseSheetLayer(t2)[0],cell=cl) for cl in data[t2][__KBS] + data[t2][__PSNR]]
                sheet.cell(row = r, column = c).value = __S_BDRATE_FORMAT.format(*(r1+r2))
                sheet.cell(row = r, column = c).style = 'Percent'
                sheet.cell(row = r, column = c).number_format = '0.00%'
            sheet.cell(row=r,column=c).alignment = xl.styles.Alignment(horizontal='center')
    # Set conditional coloring
    form_range = "{}:{}".format(get_column_letter(col)+str(row),get_column_letter(final_c)+str(final_r))
    sheet.conditional_formatting.add(form_range,
                                     ColorScaleRule(start_type='percentile', start_value=90, start_color='63BE7B',
                                                    mid_type='num', mid_value=0, mid_color='FFFFFF',
                                                    end_type='percentile', end_value=10, end_color='F8696B' ))


"""
Write bit size summary matrix
"""
def __writeBSummaryMatrix(sheet, data, row, col):
    test_col = col - 1
    test_row = row - 1
    final_r = row+len(data.keys())
    final_c = col+len(data.keys())
    for r in range(row,row+len(data.keys())):
        for c in range(col,col+len(data.keys())):
            t2 = sheet.cell(row = r, column = test_col).value
            t1 = sheet.cell(row = test_row, column = c).value
            if t1 == t2:
                sheet.cell(row = r, column = c).value = "-"
            else:
                r2 =[__SR_FORMAT.format(sheet=parseSheetLayer(t1)[0],cell=cl) for cl in data[t1][__KB]]
                r1 =[__SR_FORMAT.format(sheet=parseSheetLayer(t2)[0],cell=cl) for cl in data[t2][__KB]]
                sheet.cell(row = r, column = c).value = __S_BIT_FORMAT.format(*(r1+r2))
                sheet.cell(row = r, column = c).style = 'Percent'
            sheet.cell(row=r,column=c).alignment = xl.styles.Alignment(horizontal='center')
    # Set conditional coloring
    form_range = "{}:{}".format(get_column_letter(col)+str(row),get_column_letter(final_c)+str(final_r))
    sheet.conditional_formatting.add(form_range,
                                     ColorScaleRule(start_type='min', start_color='4F81BD',
                                                    mid_type='num', mid_value=1, mid_color='FFFFFF',
                                                    end_type='percentile', end_value=80, end_color='F8696B' ))

"""
Write time summary matrix
"""
def __writeTimeSummaryMatrix(sheet, data, row, col):
    test_col = col - 1
    test_row = row - 1
    final_r = row+len(data.keys())
    final_c = col+len(data.keys())
    for r in range(row,row+len(data.keys())):
        for c in range(col,col+len(data.keys())):
            t2 = sheet.cell(row = r, column = test_col).value
            t1 = sheet.cell(row = test_row, column = c).value
            if t1 == t2:
                sheet.cell(row = r, column = c).value = "-"
            else:
                r2 =[__SR_FORMAT.format(sheet=parseSheetLayer(t1)[0],cell=cl) for cl in data[t1][__TIME]]
                r1 =[__SR_FORMAT.format(sheet=parseSheetLayer(t2)[0],cell=cl) for cl in data[t2][__TIME]]
                sheet.cell(row = r, column = c).value = __S_TIME_FORMAT.format(*(r1+r2))
                sheet.cell(row = r, column = c).style = 'Percent'
            sheet.cell(row=r,column=c).alignment = xl.styles.Alignment(horizontal='center')
    # Set conditional coloring
    form_range = "{}:{}".format(get_column_letter(col)+str(row),get_column_letter(final_c)+str(final_r))
    sheet.conditional_formatting.add(form_range,
                                     ColorScaleRule(start_type='min', start_color='9BDE55',#'63BE7B',
                                                    mid_type='num', mid_value=1, mid_color='FFFFFF',
                                                    end_type='percentile', end_value=80, end_color='00BBEF'))#'4F81BD' ))

"""
Write psnr summary matrix
"""
def __writePSNRSummaryMatrix(sheet, data, row, col):
    test_col = col - 1
    test_row = row - 1
    final_r = row+len(data.keys())
    final_c = col+len(data.keys())
    for r in range(row,row+len(data.keys())):
        for c in range(col,col+len(data.keys())):
            t2 = sheet.cell(row = r, column = test_col).value
            t1 = sheet.cell(row = test_row, column = c).value
            if t1 == t2:
                sheet.cell(row = r, column = c).value = 0#"-"
            else:
                r2 =[__SR_FORMAT.format(sheet=parseSheetLayer(t1)[0],cell=cl) for cl in data[t1][__PSNR]]
                r1 =[__SR_FORMAT.format(sheet=parseSheetLayer(t2)[0],cell=cl) for cl in data[t2][__PSNR]]
                sheet.cell(row = r, column = c).value = __S_PSNR_FORMAT.format(*(r1+r2))
            sheet.cell(row = r, column = c).style = 'Comma'
            sheet.cell(row=r,column=c).alignment = xl.styles.Alignment(horizontal='center')
    # Set conditional coloring
    form_range = "{}:{}".format(get_column_letter(col)+str(row),get_column_letter(final_c)+str(final_r))
    sheet.conditional_formatting.add(form_range,
                                     ColorScaleRule(start_type='min', start_color='FF772A',
                                                    mid_type='num', mid_value=0, mid_color='FFFFFF',
                                                    end_type='max', end_color='9BDE55' ))

"""
Write summary sheet
"""
def __writeSummary(sheet, ref_pos, order = None):

    seq_ref = {}
    order = order if order else list(seq_ref.keys())

    # Init structure for each sequence
    for (test,item) in ref_pos.items():
        for seq in item.keys():
            seq_ref[seq] = {}
    # Populate
    for (test,item) in ref_pos.items():
        for (seq,val) in item.items():
            seq_ref[seq][test] = val
    #print(seq_ref)
    # For each sequence generate the comparison matrix
    sheet.cell(row = 1, column = 1).value = __S_HEADER 
    #for (seq,ref) in sorted(seq_ref.items()):
    for seq in order:
        ref = seq_ref[seq]
        tests = sorted(ref.keys())
        
        # write bdrate matrix
        row = sheet.max_row + 2
        brow = row
        prow = row
        trow = row
        col = 1 #sheet.max_column + 1
        sheet.cell(row = row, column = col).value = __S_SEQ_HEADER.format(seq)
        sheet.merge_cells(start_column=col,start_row=row,end_column=col+len(tests),end_row=row)
        (row, col) = __writeSummaryMatrixHeader(sheet, tests, row+1, col)
        __writeBDSummaryMatrix(sheet, ref, row, col)

        # write bit matrix
        if 'bcol' not in locals():
            bcol = sheet.max_column + 2
        sheet.cell(row = brow, column = bcol).value = __S_BIT_HEADER
        sheet.merge_cells(start_column=bcol,start_row=brow,end_column=bcol+len(tests),end_row=brow)
        (brow, col) = __writeSummaryMatrixHeader(sheet, tests, brow+1, bcol)
        __writeBSummaryMatrix(sheet, ref, brow, col)

        # write psnr matrix
        if 'pcol' not in locals():
            pcol = sheet.max_column + 2
        sheet.cell(row = prow, column = pcol).value = __S_PSNR_HEADER
        sheet.merge_cells(start_column=pcol,start_row=prow,end_column=pcol+len(tests),end_row=prow)
        (prow, col) = __writeSummaryMatrixHeader(sheet, tests, prow+1, pcol)
        __writePSNRSummaryMatrix(sheet, ref, prow, col)

        # write time matrix
        if 'tcol' not in locals():
            tcol = sheet.max_column + 2
        sheet.cell(row = trow, column = tcol).value = __S_TIME_HEADER
        sheet.merge_cells(start_column=tcol,start_row=trow,end_column=tcol+len(tests),end_row=trow)
        (trow, col) = __writeSummaryMatrixHeader(sheet, tests, trow+1, tcol)
        __writeTimeSummaryMatrix(sheet, ref, trow, col)

    # Make columns wider
    for col in range(sheet.max_column):
        sheet.column_dimensions[get_column_letter(col+1)].width = getMaxLength(list(ref_pos.keys()))

"""
Write summary 2 list header
"""
def __writeSummary2ListHeader(sheet, tests, seqs, base_test, row, col):
    # Write the test names in the first rows
    base_test_col = -1;

    tmp_col = col + 1
    for test in tests:
        sheet.cell(row = row, column = tmp_col).value = test
        if test == base_test:
            base_test_col = tmp_col
        tmp_col += 1

    #Write vertical 
    tmp_row = row + 1
    for seq in seqs:
        sheet.cell(row = tmp_row, column = col).value = seq
        tmp_row += 1

    return base_test_col

"""
Write summary 2 list
"""
def __writeSummary2List(sheet, data, base_test_col, row, col):
    seq_col = col - 1
    test_row = row - 1
    final_r = row+len(data.keys())
    final_c = col+len(tuple(data.values())[0].keys())
    for r in range(row,final_r):
        for c in range(col,final_c):
            seq = sheet.cell(row = r, column = seq_col).value
            test = sheet.cell(row = test_row, column = c).value
            base_test = sheet.cell(row = test_row, column = base_test_col).value
            #if t1 == t2:
            #    sheet.cell(row = r, column = c).value = "-"
            #else:
            base =[__SR_FORMAT.format(sheet=parseSheetLayer(base_test)[0],cell=cl) for cl in data[seq][base_test][__KBS] + data[seq][base_test][__PSNR]]
            comp =[__SR_FORMAT.format(sheet=parseSheetLayer(test)[0],cell=cl) for cl in data[seq][test][__KBS] + data[seq][test][__PSNR]]
            sheet.cell(row = r, column = c).value = __S_BDRATE_FORMAT.format(*(base + comp))
            sheet.cell(row = r, column = c).style = 'Percent'
            sheet.cell(row=r,column=c).alignment = xl.styles.Alignment(horizontal='center')
    # Set conditional coloring
    form_range = "{}:{}".format(get_column_letter(col)+str(row),get_column_letter(final_c)+str(final_r))
    sheet.conditional_formatting.add(form_range,
                                     ColorScaleRule(start_type='percentile', start_value=90, start_color='63BE7B',
                                                    mid_type='num', mid_value=0, mid_color='FFFFFF',
                                                    end_type='percentile', end_value=10, end_color='F8696B' ))


"""
Write summary sheet 2
Write a list type summary with one reference point
"""
def __writeSummary2(sheet, ref_pos, base_test, order=None):

    seq_ref = {}

    # Init structure for each sequence
    for (test,item) in ref_pos.items():
        for seq in item.keys():
            seq_ref[seq] = {}
    # Populate
    for (test,item) in ref_pos.items():
        for (seq,val) in item.items():
            seq_ref[seq][test] = val
    #print(seq_ref)

    order = order if order else sorted(seq_ref.keys())

    # List Each sequence (rows) and tests (columns)
    sheet.cell(row = 1, column = 1).value = __S2_HEADER
    row = sheet.max_row + 2
    col = 1

    base_test_col = __writeSummary2ListHeader(sheet, sorted(tuple(seq_ref.values())[0].keys()), order, base_test, row, col)
    __writeSummary2List(sheet, seq_ref, base_test_col, row + 1, col + 1)
   

    # Make columns wider
    for col in range(sheet.max_column):
        sheet.column_dimensions[get_column_letter(col+1)].width = getMaxLength(list(ref_pos.keys()))

"""
Transform res_pos into summary test structure
"""
def __makeSummary(res_pos,layers={}):
    res = {}
    for (test,item) in res_pos.items():
        res[test] = {}
        for (seq,vals) in item.items():
            test_in_layers = test in layers
            if not test_in_layers:
                res[test][seq] = vals[__LID_TOT]
                if len(vals.keys()) <= 2: #If only 2 lids, it should mean the total layer and other layer are the same
                    continue
            for (lid,val) in vals.items():
                if lid == __LID_TOT and not test_in_layers:
                    continue
                if test_in_layers and lid not in layers[test]:
                    continue
                nn = makeSheetLayer(test,lid)
                if lid == __LID_TOT or len(vals.keys()) <= 2:
                    nn = test
                if nn in res:
                    res[nn][seq] = val
                else:
                    res[nn] = {seq:val}

    return res

"""
Write results for a single test/sheet
@return positions of relevant cells
"""
def __writeSheet(sheet,data,scale,qp_names,order=None):
    # Write header
    for col in range(len(__R_HEADER)):
        sheet.cell(row = 1, column = col+1).value = __R_HEADER[col]
        sheet.merge_cells(start_column=col+1,start_row=1,end_column=col+1,end_row=3)
        sheet.cell(row=1,column=col+1).alignment = xl.styles.Alignment(horizontal='center')

    sheet.column_dimensions['A'].width = 50

    qp_cols = {}
    seq_rows = {}
    res_ref = {}

    order = order if order else list(data.keys())

    # Write stat row
    for (qp,item) in sorted(tuple(data.values())[0].items()):
        qp_cols[qp] = sheet.max_column+1
        for val in __R_KBS:
            sheet.cell(row=2,column=sheet.max_column+1).value = val
            sheet.merge_cells(start_column=sheet.max_column,start_row=2,end_column=sheet.max_column,end_row=3)
            sheet.cell(row=2,column=sheet.max_column).alignment = xl.styles.Alignment(horizontal='center')

        for val in __R_PSNR_SUB:
            sheet.cell(row=3,column=sheet.max_column+1).value = val
        sheet.cell(row=2,column=sheet.max_column-len(__R_PSNR_SUB)+1).value = __R_PSNR
        sheet.merge_cells(start_column=sheet.max_column-len(__R_PSNR_SUB)+1,start_row=2,end_column=sheet.max_column,end_row=2)
        sheet.cell(row=2,column=sheet.max_column-len(__R_PSNR_SUB)+1).alignment = xl.styles.Alignment(horizontal='center')
        
        sheet.cell(row=1,column=qp_cols[qp]).value = __R_HEADER_QP.format(tuple(qp_names.values())[0][qp])
        sheet.merge_cells(start_column=qp_cols[qp],start_row=1,end_column=sheet.max_column,end_row=1)
        sheet.cell(row=1,column=qp_cols[qp]).alignment = xl.styles.Alignment(horizontal='center')

    # Write sequence column
    #for (seq,res) in data.items():
    layer_r = range(len(tuple(tuple(data.values())[0].values())[0].keys())-1)
    for seq in order:
        sheet.cell(row=sheet.max_row+1,column=1).value = __SEQ_FORMAT.format(seq,scale) if seq is not __SEQ_AVERAGE else __SEQ_AVERAGE 
        seq_rows[seq] = sheet.max_row

        res_ref[seq] = {}
            
        #Set Layers
        sheet.cell(row=seq_rows[seq],column=2).value = __LID_TOT
        res_ref[seq][__LID_TOT] = {__KB:[], __KBS:[],__PSNR:[],__TIME:[]}
        for lid in layer_r:
            sheet.cell(row=seq_rows[seq]+lid+1,column=2).value = lid
            res_ref[seq][lid] = {__KB:[], __KBS:[],__PSNR:[],__TIME:[]}
        

    # Write Average "sequence"
    #sheet.cell(row=sheet.max_row+1,column=1).value = __SEQ_AVERAGE
    #seq_rows[__SEQ_AVERAGE] = sheet.max_row

    #res_ref[__SEQ_AVERAGE] = {}
            
    ##Set Layers
    #sheet.cell(row=seq_rows[__SEQ_AVERAGE],column=2).value = __LID_TOT
    #res_ref[__SEQ_AVERAGE][__LID_TOT] = {__KB:[], __KBS:[],__PSNR:[],__TIME:[]}
    #for lid in range(len(tuple(res.values())[0].keys())-1):
    #    sheet.cell(row=seq_rows[__SEQ_AVERAGE]+lid+1,column=2).value = lid
    #    res_ref[__SEQ_AVERAGE][lid] = {__KB:[], __KBS:[],__PSNR:[],__TIME:[]}
    

    # Set actual data
    #for (seq,qps) in data.items():
    for seq in order:
        if seq in data:
            qps = data[seq]
            for (qp,item) in sorted(qps.items()):
                for (lid,val) in item.items():
                    r = seq_rows[seq] + lid + 1
                    if lid == __LID_TOT:
                        r = seq_rows[seq]
                    c_kb = qp_cols[qp]
                    c_kbs = c_kb + 1
                    c_time = c_kbs + 1
                    c_psnr = c_time + len(__R_PSNR_SUB)

                    sheet.cell(row=r,column=c_kb).value = val[__KB]
                    sheet.cell(row=r,column=c_kbs).value = val[__KBS]
                    sheet.cell(row=r,column=c_time).value = val[__TIME]

                    for i in range(len(__R_PSNR_SUB)-1):
                        sheet.cell(row=r,column=c_psnr-i-1).value = float(val[__PSNR][-i-1])
                    sheet.cell(row=r,column=c_psnr).value = __PSNR_AVG.format(y = get_column_letter(c_psnr-3) + str(r),
                                                                              u = get_column_letter(c_psnr-2) + str(r),
                                                                              v = get_column_letter(c_psnr-1) + str(r))

                    res_ref[seq][lid][__KB].append(get_column_letter(c_kb) + str(r))
                    res_ref[seq][lid][__KBS].append(get_column_letter(c_kbs) + str(r))
                    res_ref[seq][lid][__TIME].append(get_column_letter(c_time) + str(r))
                    res_ref[seq][lid][__PSNR].append(get_column_letter(c_psnr) + str(r))
        else:
            # Set Average data
            for c_kb in sorted(qp_cols.values()):
                for lid in res_ref[__SEQ_AVERAGE]:
                    r = seq_rows[__SEQ_AVERAGE] + lid + 1
                    if lid == __LID_TOT:
                        r = seq_rows[__SEQ_AVERAGE]
                    c_kbs = c_kb + 1
                    c_time = c_kbs + 1
                    c_psnr = c_time + len(__R_PSNR_SUB)

                    kb_rows = []
                    kbs_rows = []
                    time_rows = []
                    for (seq,row) in seq_rows.items():
                        if seq == __SEQ_AVERAGE:
                            continue
                        kb_rows.append(get_column_letter(c_kb)+str(row+lid+1))
                        kbs_rows.append(get_column_letter(c_kbs)+str(row+lid+1))
                        time_rows.append(get_column_letter(c_time)+str(row+lid+1))
                    sheet.cell(row=r,column=c_kb).value = __C_AVG.format(','.join(kb_rows))
                    sheet.cell(row=r,column=c_kbs).value = __C_AVG.format(','.join(kbs_rows))
                    sheet.cell(row=r,column=c_time).value = __C_AVG.format(','.join(time_rows))

                    for i in range(len(__R_PSNR_SUB)-1):
                        psnr_rows = []
                        for (seq,row) in seq_rows.items():
                            if seq == __SEQ_AVERAGE:
                                continue
                            psnr_rows.append(get_column_letter(c_psnr-i-1)+str(row+lid+1))
                        sheet.cell(row=r,column=c_psnr-i-1).value = __C_AVG.format(','.join(psnr_rows))
            
                    sheet.cell(row=r,column=c_psnr).value = __PSNR_AVG.format(y = get_column_letter(c_psnr-3) + str(r),
                                                                              u = get_column_letter(c_psnr-2) + str(r),
                                                                              v = get_column_letter(c_psnr-1) + str(r))

                    res_ref[__SEQ_AVERAGE][lid][__KB].append(get_column_letter(c_kb) + str(r))
                    res_ref[__SEQ_AVERAGE][lid][__KBS].append(get_column_letter(c_kbs) + str(r))
                    res_ref[__SEQ_AVERAGE][lid][__TIME].append(get_column_letter(c_time) + str(r))
                    res_ref[__SEQ_AVERAGE][lid][__PSNR].append(get_column_letter(c_psnr) + str(r))


    sheet.freeze_panes = 'A4'

    return res_ref


"""
Write results to workbook. Assume wb has a Summary and Result Sheet.
"""
def __writeResults(wb,results,layers={},s2_base=None):
    s_sheet = wb.get_sheet_by_name('Summary')
    s2_sheet = wb.get_sheet_by_name('Summary2')
    res_pos = {}
    
    # Write test results
    for (test,res) in sorted(results.items()):
        n_sheet = wb.create_sheet(title=test,index=0)
        res_pos[test] = __writeSheet(n_sheet,res[__RES],res[__SCALE],res[__QPS],res[__INAMES])
    res_pos = __makeSummary(res_pos,layers)
    #print(res_pos)
    #write summary sheet
    __writeSummary(s_sheet,res_pos,res[__INAMES])
    __writeSummary2(s2_sheet,res_pos,s2_base if s2_base else tuple(sorted(res_pos.keys()))[-1],res[__INAMES])
    wb.active = wb.get_index(s_sheet)

"""
Run given tests and write results to a exel file
@param combi: Give a list of test names that are combined into one test
@param layer_combi: Given tests are combined as like they were layers
@param layers: A dict with test names as keys containing a list of layers to include in summary
@param s2_base: Test name of the s2 summary that should be the base of the comparison
"""
def runTests( tests, outname, combi = [], layer_combi = [], layers = {}, s2_base = None, input_res = False ):
    print('Start running tests...')
    nt = 1
    for test in tests:
        #print("Running test {}...".format(test._test_name))
        print_out = "[{}/{}] ".format(nt,len(tests))
        print(print_out, end='\r')
        test.run(print_out, input_res)
        nt += 1
    print('Tests complete.')
    print('Writing results to file {}...'.format(cfg.results + outname + __FILE_END))
    res = __parseTestResults(tests)
    res = __combiTestResults(res,combi,layer_combi)

    wb = xl.load_workbook(cfg.exel_template,keep_vba=True)
    __writeResults(wb,res,layers,s2_base)
    wb.save(cfg.results + outname + __FILE_END)
    print('Done.')