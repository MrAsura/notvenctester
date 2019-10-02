"""
Module for defining and processing summary sheets
"""

import openpyxl as xl

#Define summary names used as keys f in definitions
sn_BDBRM = "BDBRMatrix"
sn_ANCHOR = "AnchorList"

#Define definition templates for each summary type

"""
__LAYERS:{<test_name>:<tuple of layers (lid) to include>}
"""
__LAYERS = "layers"
__WRITE_BITS = "write_bits"
__WRITE_BDBR = "write_bdbr"
__WRITE_PSNR = "write_psrn"
__WRITE_TIME = "write_time"
dt_BDBRM = {__LAYERS:{}, __WRITE_BDBR: True, __WRITE_BITS: True, __WRITE_PSNR: True, __WRITE_TIME: True}

"""
for creating the BDBRMatrix definition
"""
def create_BDBRMatrix_definition(layers, write_bdbr, write_bits, write_psnr, write_time):
    definition = dt_BDBRM.copy()
    definition[__LAYERS] = layers
    definition[__WRITE_BDBR] = write_bdbr
    definition[__WRITE_BITS] = write_bits
    definition[__WRITE_PSNR] = write_psnr
    definition[__WRITE_TIME] = write_time
    return definition

"""
__????:dt_ANCHOR_SUB|None
"""
__BDBR = "bdbr"
__BITS = "bits"
__PSNR = "psnr"
__TIME = "time"
dt_ANCHOR = {__BDBR:{}, __BITS:{}, __PSNR:{}, __TIME:{}}
"""
dt_ANCHOR_SUB = {<test_name>:(<anchor_test_name>|None,...)}
"""

"""
Create anchor list definition. Pass in sub definitions for each data type that should be included
"""
def create_AnchorList_definition(bdbr_def, bits_def, psnr_def, time_def):
    definition = dt_ANCHOR.copy()
    definition[__BDBR] = create_AnchorSub_definition(bdbr_def)
    definition[__BITS] = create_AnchorSub_definition(bits_def)
    definition[__PSNR] = create_AnchorSub_definition(psnr_def)
    definition[__TIME] = create_AnchorSub_definition(time_def)
    return definition

"""
Sub anchor list definition. Takes in a list of test anchor pairs with optional layer ids ((<name>,#lid):(<anchor>,#lid))
"""
def create_AnchorSub_definition(definition):
    from TestSuite import makeSheetLayer, _LID_TOT
    def processTestName(tn):
        if not tn:
            return None
        if isinstance(tn, str):
            return tn
        if len(tn) == 2 and isinstance(tn, tuple):
            if isinstance(tn[0], str) and isinstance(tn[1], int):
                return makeSheetLayer(tn[0], tn[1]) if tn[1] != _LID_TOT else tn[0]
        return tuple(processTestName(test) for test in tn)

    if not definition:
        return None
    if isinstance(definition, tuple):
        return {processTestName(definition[0]): processTestName(definition[1])}
    if isinstance(definition, dict):
        definition = definition.items()
    return {processTestName(name): processTestName(val) for (name, val) in definition if val}

"""
Function for creating summary sheets to the given workbook for the given data
@param wb: work book where sheets will be added
@param data_refs: references to the data sheets and data cell positions used for the summary in the form data_refs[<test_name>][<seq>][<lid>] = {__KB, __KBS,__PSNR, __TIME}
@param BDBRMatrix: dict containing the parameters for the BDBRMatrix summary type
"""
def makeSummaries(wb, data_refs, order = None, **definitions):
    if sn_BDBRM in definitions:
        makeBDBRMatrix(wb, data_refs, order, definitions[sn_BDBRM])
    if sn_ANCHOR in definitions:
        makeAnchorList(wb, data_refs, order, definitions[sn_ANCHOR])


def makeBDBRMatrix(wb, data_refs, order, definition):
    bdbrm_sheet = wb.create_sheet(sn_BDBRM)
    expanded_refs = __makeSummary(data_refs, definition[__LAYERS])
    __writeBDBRMatrix(bdbrm_sheet, expanded_refs, order, **definition)

def makeAnchorList(wb, data_refs, order, definition):
    anchor_sheet = wb.create_sheet(sn_ANCHOR)
    expanded_refs = __makeSummary(data_refs)
    __writeAnchorList(anchor_sheet, expanded_refs, order, **definition)

"""
Transform res_pos into summary test structure making test layers into their own tests
@return dict of the form res[<test_name>][<seq>] = {__KB, __KBS,__PSNR, __TIME}
"""
def __makeSummary(res_pos,layers={}):
    from TestSuite import _LID_TOT, makeSheetLayer
    res = {}
    for (test,item) in res_pos.items():
        res[test] = {}
        for (seq,vals) in item.items():
            test_in_layers = test in layers
            if not test_in_layers:
                res[test][seq] = vals[LID_TOT]
                if len(vals.keys()) <= 2: #If only 2 lids, it should mean the total layer and other layer are the same
                    continue
            for (lid,val) in vals.items():
                if lid == LID_TOT and not test_in_layers:
                    continue
                if test_in_layers and lid not in layers[test]:
                    continue
                nn = makeSheetLayer(test,lid)
                if lid == LID_TOT or len(vals.keys()) <= 2:
                    nn = test
                if nn in res:
                    res[nn][seq] = val
                else:
                    res[nn] = {seq:val}

    return res

"""
Switch the first and second level keys of the dict
"""
def __flip_dict(old_d):
    new_d = {val_key: {} for val in old_d.values() for val_key in val.keys()}
    for (key, val) in old_d.items():
        for (val_key, sub_val) in val.items():
            new_d[val_key][key] = sub_val
    return new_d


def getMaxLength( string_list ):
    return max((len(x) for x in string_list))


__S_BDRATE_FORMAT = "=bdrate({},{},{},{},{},{},{},{},{},{},{},{},{},{},{},{})"
__S_BIT_FORMAT = "=AVERAGE({},{},{},{})/AVERAGE({},{},{},{})"
__S_TIME_FORMAT = "=AVERAGE({},{},{},{})/AVERAGE({},{},{},{})"
__S_PSNR_FORMAT = "=AVERAGE({},{},{},{})-AVERAGE({},{},{},{})"
__SR_FORMAT = r"'{sheet}'!{cell}"


#######################################
# AnchorList summary type definitions #
#######################################
__AL_HEADER = r"Result anchor list summary"
__AL_SEQ_FORMAT = r"{}"
__AL_BD_HEADER = r"BDBR results"
__AL_B_HEADER = r"Bit results"
__AL_PSNR_HEADER = r"PSNR results"
__AL_TIME_HEADER = r"Time results"
__AL_TEST = r"Test:"
__AL_ANCHOR = r"Anchor:"
__AL_SEQ = r"Sequences:"

"""
Handle writing the anchor list structure
"""

def __writeAnchorList(sheet, data_refs, order = None, *, bdbr, bits, psnr, time, **other):
    seq_ref = __flip_dict(data_refs) # transform data_refs to seq_ref[<seq>][<test_name>] order
    order = order if order else list(seq_ref.keys())

    # Each sequence is one line so generate one columns for each test in each gategory based on the given definitions
    sheet.cell(row = 1, column = 1).value = __AL_HEADER 
    
    for seq in order:

        row = sheet.max_row + 1
        col = 3 #sheet.max_column + 1
        first_row = row + 1

        # write bdrate tests
        if bdbr:
            if 'bdcol' not in locals():
                bdcol = sheet.max_column + 2
                sheet.cell(row = row-1, column = bdcol - 1).value = __AL_BD_HEADER
                sheet.cell(row = row, column = bdcol - 1).value = __AL_TEST
                sheet.cell(row = row+1, column = bdcol - 1).value = __AL_SEQ
                __writeAnchorListHeader(sheet, bdbr, row, bdcol)
                row += 1

            # Write sequence
            sheet.cell(row = row, column = col).value = __AL_SEQ_FORMAT.format(seq)
            sheet.cell(row = row, column = col - 1).value = __AL_SEQ

            __writeAnchorListData(sheet, ref[seq], bdbr, row, bdcol,
                                     data_func = lambda data, test: data[test][_KBS] + data[test][_PSNR],
                                     data_format = __S_BDRATE_FORMAT,
                                     number_format = '0.00%',
                                     color_scale_rule = ColorScaleRule(start_type='percentile', start_value=90, start_color='63BE7B',
                                                                       mid_type='num', mid_value=0, mid_color='FFFFFF',
                                                                       end_type='percentile', end_value=10, end_color='F8696B' ))
                                   

        # write bit tests
        if bits:
            if 'bcol' not in locals():
                bcol = sheet.max_column + 2
                sheet.cell(row = row-1, column = bcol - 1).value = __AL_B_HEADER
                sheet.cell(row = row, column = bcol - 1).value = __AL_TEST
                sheet.cell(row = row+1, column = bcol - 1).value = __AL_SEQ
                __writeAnchorListHeader(sheet, bits, row, bcol)
                row += 1

            # Write sequence
            sheet.cell(row = row, column = col).value = __AL_SEQ_FORMAT.format(seq)
            sheet.cell(row = row, column = col - 1).value = __AL_SEQ

            __writeAnchorListData(sheet, ref[seq], bits, row, bcol,
                                     data_func = lambda data, test: data[test][_KB],
                                     data_format = __S_BIT_FORMAT,
                                     color_scale_rule = ColorScaleRule(start_type='min', start_color='4F81BD',
                                                                       mid_type='num', mid_value=1, mid_color='FFFFFF',
                                                                       end_type='percentile', end_value=80, end_color='F8696B' ))

        # write psnr tests
        if psnr:
            if 'pcol' not in locals():
                pcol = sheet.max_column + 2
                sheet.cell(row = row-1, column = pcol - 1).value = __AL_B_HEADER
                sheet.cell(row = row, column = pcol - 1).value = __AL_TEST
                sheet.cell(row = row+1, column = pcol - 1).value = __AL_SEQ
                __writeAnchorListHeader(sheet, psnr, row, pcol)
                row += 1

            # Write sequence
            sheet.cell(row = row, column = col).value = __AL_SEQ_FORMAT.format(seq)
            sheet.cell(row = row, column = col - 1).value = __AL_SEQ

            __writeAnchorListData(sheet, ref[seq], psnr, row, pcol,
                                     data_func = lambda data, test: data[test][_PSNR],
                                     data_format = __S_PSNR_FORMAT,
                                     data_style = 'Comma',
                                     def_val = 0,
                                     color_scale_rule = ColorScaleRule(start_type='percentile', start_value=90, start_color='63BE7B',
                                                                       mid_type='num', mid_value=0, mid_color='FFFFFF',
                                                                       end_type='percentile', end_value=10, end_color='F8696B' ))

        # write time matrix
        if time:
            if 'tcol' not in locals():
                tcol = sheet.max_column + 2
                sheet.cell(row = row-1, column = tcol - 1).value = __AL_B_HEADER
                sheet.cell(row = row, column = tcol - 1).value = __AL_TEST
                sheet.cell(row = row+1, column = tcol - 1).value = __AL_SEQ
                __writeAnchorListHeader(sheet, time, row, tcol)
                row += 1

            # Write sequence
            sheet.cell(row = row, column = col).value = __AL_SEQ_FORMAT.format(seq)
            sheet.cell(row = row, column = col - 1).value = __AL_SEQ

            __writeAnchorListData(sheet, ref[seq], time, row, tcol,
                                     data_func = lambda data, test: data[test][_TIME],
                                     data_format = __S_TIME_FORMAT,
                                     color_scale_rule = ColorScaleRule(start_type='min', start_color='9BDE55',#'63BE7B',
                                                                       mid_type='num', mid_value=1, mid_color='FFFFFF',
                                                                       end_type='percentile', end_value=80, end_color='00BBEF'))

    # Make columns wider
    for column in range(sheet.max_column):
        sheet.column_dimensions[get_column_letter(column+1)].width = getMaxLength(list(data_refs.keys()) + list(order))

    #Add conditional formatting
    form_ranges = []
    color_rules = []
    #BDRATE
    if bdbr:
        form_ranges.append("{}:{}".format(get_column_letter(bdcol)+str(first_row),get_column_letter(len(bdbr.keys()))+str(row)))
        color_rules.append(ColorScaleRule(start_type='percentile', start_value=90, start_color='63BE7B',
                                          mid_type='num', mid_value=0, mid_color='FFFFFF',
                                          end_type='percentile', end_value=10, end_color='F8696B' ))
    if bits:
        form_ranges.append("{}:{}".format(get_column_letter(bcol)+str(first_row),get_column_letter(len(bits.keys()))+str(row)))
        color_rules.append(ColorScaleRule(start_type='min', start_color='4F81BD',
                                          mid_type='num', mid_value=1, mid_color='FFFFFF',
                                          end_type='percentile', end_value=80, end_color='F8696B' ))
    if psnr:
        form_ranges.append("{}:{}".format(get_column_letter(pcol)+str(first_row),get_column_letter(len(psnr.keys()))+str(row)))
        color_rules.append(ColorScaleRule(start_type='percentile', start_value=90, start_color='63BE7B',
                                          mid_type='num', mid_value=0, mid_color='FFFFFF',
                                          end_type='percentile', end_value=10, end_color='F8696B' ))
    if time:
        form_ranges.append("{}:{}".format(get_column_letter(tcol)+str(first_row),get_column_letter(len(time.keys()))+str(row)))
        color_rules.append(ColorScaleRule(start_type='min', start_color='9BDE55',#'63BE7B',
                                          mid_type='num', mid_value=1, mid_color='FFFFFF',
                                          end_type='percentile', end_value=80, end_color='00BBEF'))

    for (f_range, c_rule) in zip(form_range, color_rules):
        sheet.conditional_formatting.add(f_range, c_rule)


def __writeAnchorListHeader(sheet, sub_def, row, col):
    #Write horizontal headers/test names
    tmp_col = col
    for (test,anchors) in sub_def.items():
        for anchor in anchors:
            sheet.cell(row = row, column = tmp_col).value = test
            sheet.cell(row = row + 1, column = tmp_col).value = anchor
            tmp_col += 1


def __writeAnchorListData(sheet, ref, sub_def, row, col, *, data_func, data_format, number_format = None, number_style = 'Percent'):
    #final_r = row+len(data.keys())
    #final_c = col+len(data.keys())
    for (c, (test, anchors)) in zip(range(col,col+len(sub_def.keys())), sub_def.items()):
        for anchor in anchors:
            test_res =[__SR_FORMAT.format(sheet=parseSheetLayer(test)[0],cell=cl) for cl in data_func(ref, test)]
            anchor_res =[__SR_FORMAT.format(sheet=parseSheetLayer(anchor)[0],cell=cl) for cl in data_func(ref, anchor)]
            sheet.cell(row = row, column = c).value = data_format.format(*(test_res+anchor_res))
            sheet.cell(row = row, column = c).style = number_style
            if number_format:
                sheet.cell(row = row, column = c).number_format = number_format
            sheet.cell(row=row,column=c).alignment = xl.styles.Alignment(horizontal='center')
    # Set conditional coloring
    #form_range = "{}:{}".format(get_column_letter(col)+str(row),get_column_letter(final_c)+str(final_r))
    #sheet.conditional_formatting.add(form_range, color_scale_rule)


#######################################
# BDBRMatrix summary type definitions #
#######################################

__S_SEQ_HEADER = "Sequence {} results"
__S_BIT_HEADER = r"Bit comparisons"
__S_PSNR_HEADER = r"PSNR comparisons (dB)"
__S_TIME_HEADER = r"Encoding time comparisons"
__S_HEADER = "Result summary matrix (bdrate, bit, PSNR, Time comparisons)"


"""
Handle writing the BDBRMatrix summary sheet
"""
def __writeBDBRMatrix(sheet, data_refs, order = None, *, write_bdbr, write_bits, write_psnr, write_time, **other):
    from TestSuite import _PSNR, _KBS, _KB, _TIME
    
    seq_ref = __flip_dict(data_refs) # transform data_refs to seq_ref[<seq>][<test_name>] order
    order = order if order else list(seq_ref.keys())

    #print(seq_ref)
    # For each sequence generate the comparison matrix
    sheet.cell(row = 1, column = 1).value = __S_HEADER 
    #for (seq,ref) in sorted(seq_ref.items()):
    for seq in order:
        ref = seq_ref[seq]
        tests = sorted(ref.keys())
        
        row = sheet.max_row + 2
        brow = row
        prow = row
        trow = row
        col = 1 #sheet.max_column + 1
        
        # write bdrate matrix
        if write_bdbr:  

            sheet.cell(row = row, column = col).value = __S_SEQ_HEADER.format(seq)
            sheet.merge_cells(start_column=col,start_row=row,end_column=col+len(tests),end_row=row)
            (row, col) = __writeSummaryMatrixHeader(sheet, tests, row+1, col)
            __writeSummaryDataMatrix(sheet, ref, row, col,
                                     data_func = lambda data, test: data[test][_KBS] + data[test][_PSNR],
                                     data_format = __S_BDRATE_FORMAT,
                                     number_format = '0.00%',
                                     color_scale_rule = ColorScaleRule(start_type='percentile', start_value=90, start_color='63BE7B',
                                                                       mid_type='num', mid_value=0, mid_color='FFFFFF',
                                                                       end_type='percentile', end_value=10, end_color='F8696B' ))
                                   

        # write bit matrix
        if write_bits:
            if 'bcol' not in locals():
                bcol = sheet.max_column + 2
            sheet.cell(row = brow, column = bcol).value = __S_BIT_HEADER
            sheet.merge_cells(start_column=bcol,start_row=brow,end_column=bcol+len(tests),end_row=brow)
            (brow, col) = __writeSummaryMatrixHeader(sheet, tests, brow+1, bcol)
            __writeSummaryDataMatrix(sheet, ref, brow, colb,
                                     data_func = lambda data, test: data[test][_KB],
                                     data_format = __S_BIT_FORMAT,
                                     color_scale_rule = ColorScaleRule(start_type='min', start_color='4F81BD',
                                                                       mid_type='num', mid_value=1, mid_color='FFFFFF',
                                                                       end_type='percentile', end_value=80, end_color='F8696B' ))

        # write psnr matrix
        if write_psnr:
            if 'pcol' not in locals():
                pcol = sheet.max_column + 2
            sheet.cell(row = prow, column = pcol).value = __S_PSNR_HEADER
            sheet.merge_cells(start_column=pcol,start_row=prow,end_column=pcol+len(tests),end_row=prow)
            (prow, col) = __writeSummaryMatrixHeader(sheet, tests, prow+1, pcol)
            __writeSummaryDataMatrix(sheet, ref, prow, colb,
                                     data_func = lambda data, test: data[test][_PSNR],
                                     data_format = __S_PSNR_FORMAT,
                                     data_style = 'Comma',
                                     def_val = 0,
                                     color_scale_rule = ColorScaleRule(start_type='percentile', start_value=90, start_color='63BE7B',
                                                                       mid_type='num', mid_value=0, mid_color='FFFFFF',
                                                                       end_type='percentile', end_value=10, end_color='F8696B' ))

        # write time matrix
        if write_time:
            if 'tcol' not in locals():
                tcol = sheet.max_column + 2
            sheet.cell(row = trow, column = tcol).value = __S_TIME_HEADER
            sheet.merge_cells(start_column=tcol,start_row=trow,end_column=tcol+len(tests),end_row=trow)
            (trow, col) = __writeSummaryMatrixHeader(sheet, tests, trow+1, tcol)
            __writeSummaryDataMatrix(sheet, ref, trow, col,
                                     data_func = lambda data, test: data[test][_TIME],
                                     data_format = __S_TIME_FORMAT,
                                     color_scale_rule = ColorScaleRule(start_type='min', start_color='9BDE55',#'63BE7B',
                                                                       mid_type='num', mid_value=1, mid_color='FFFFFF',
                                                                       end_type='percentile', end_value=80, end_color='00BBEF'))

    # Make columns wider
    for col in range(sheet.max_column):
        sheet.column_dimensions[get_column_letter(col+1)].width = getMaxLength(list(data_refs.keys()))


"""
Write summary base header
@return row and col of start of data field
"""
def __writeSummaryMatrixHeader(sheet, tests, row, col):
    d_row = row + 1
    d_col = col + 1
    #Write horizontal headers/test names
    tmp_col = col + 1
    for test in tests:
        sheet.cell(row = row, column = tmp_col).value = test
        tmp_col += 1

    #Write vertical 
    tmp_row = row + 1
    for test in tests:
        sheet.cell(row = tmp_row, column = col).value = test
        tmp_row += 1

    return d_row, d_col

"""
Write summary matrix data array
"""
def __writeSummaryDataMatrix(sheet, data, row, col, *, data_func, data_format, number_format = None, number_style = 'Percent', def_val = '-', color_scale_rule):
    test_col = col - 1
    test_row = row - 1
    final_r = row+len(data.keys())
    final_c = col+len(data.keys())
    for r in range(row,row+len(data.keys())):
        for c in range(col,col+len(data.keys())):
            t2 = sheet.cell(row = r, column = test_col).value
            t1 = sheet.cell(row = test_row, column = c).value
            if t1 == t2:
                sheet.cell(row = r, column = c).value = def_val
            else:
                r1 =[__SR_FORMAT.format(sheet=parseSheetLayer(t1)[0],cell=cl) for cl in data_func(data, t1)]
                r2 =[__SR_FORMAT.format(sheet=parseSheetLayer(t2)[0],cell=cl) for cl in data_func(data, t2)]
                sheet.cell(row = r, column = c).value = data_format.format(*(r1+r2))
                sheet.cell(row = r, column = c).style = number_style
                if number_format:
                    sheet.cell(row = r, column = c).number_format = number_format
            sheet.cell(row=r,column=c).alignment = xl.styles.Alignment(horizontal='center')
    # Set conditional coloring
    form_range = "{}:{}".format(get_column_letter(col)+str(row),get_column_letter(final_c)+str(final_r))
    sheet.conditional_formatting.add(form_range, color_scale_rule)

#"""
#Write summary matrix data for bdrate
#"""
#def __writeBDSummaryMatrix(sheet, data, row, col):
#    from TestSuite import _PSNR
#    test_col = col - 1
#    test_row = row - 1
#    final_r = row+len(data.keys())
#    final_c = col+len(data.keys())
#    for r in range(row,row+len(data.keys())):
#        for c in range(col,col+len(data.keys())):
#            t2 = sheet.cell(row = r, column = test_col).value
#            t1 = sheet.cell(row = test_row, column = c).value
#            if t1 == t2:
#                sheet.cell(row = r, column = c).value = "-"
#            else:
#                r1 =[__SR_FORMAT.format(sheet=parseSheetLayer(t1)[0],cell=cl) for cl in data[t1][_KBS] + data[t1][_PSNR]]
#                r2 =[__SR_FORMAT.format(sheet=parseSheetLayer(t2)[0],cell=cl) for cl in data[t2][_KBS] + data[t2][_PSNR]]
#                sheet.cell(row = r, column = c).value = __S_BDRATE_FORMAT.format(*(r1+r2))
#                sheet.cell(row = r, column = c).style = 'Percent'
#                sheet.cell(row = r, column = c).number_format = '0.00%'
#            sheet.cell(row=r,column=c).alignment = xl.styles.Alignment(horizontal='center')
#    # Set conditional coloring
#    form_range = "{}:{}".format(get_column_letter(col)+str(row),get_column_letter(final_c)+str(final_r))
#    sheet.conditional_formatting.add(form_range,
#                                     ColorScaleRule(start_type='percentile', start_value=90, start_color='63BE7B',
#                                                    mid_type='num', mid_value=0, mid_color='FFFFFF',
#                                                    end_type='percentile', end_value=10, end_color='F8696B' ))


#"""
#Write bit size summary matrix
#"""
#def __writeBSummaryMatrix(sheet, data, row, col):
#    from TestSuite import _KB
#    test_col = col - 1
#    test_row = row - 1
#    final_r = row+len(data.keys())
#    final_c = col+len(data.keys())
#    for r in range(row,row+len(data.keys())):
#        for c in range(col,col+len(data.keys())):
#            t2 = sheet.cell(row = r, column = test_col).value
#            t1 = sheet.cell(row = test_row, column = c).value
#            if t1 == t2:
#                sheet.cell(row = r, column = c).value = "-"
#            else:
#                r2 =[__SR_FORMAT.format(sheet=parseSheetLayer(t1)[0],cell=cl) for cl in data[t1][_KB]]
#                r1 =[__SR_FORMAT.format(sheet=parseSheetLayer(t2)[0],cell=cl) for cl in data[t2][_KB]]
#                sheet.cell(row = r, column = c).value = __S_BIT_FORMAT.format(*(r1+r2))
#                sheet.cell(row = r, column = c).style = 'Percent'
#            sheet.cell(row=r,column=c).alignment = xl.styles.Alignment(horizontal='center')
#    # Set conditional coloring
#    form_range = "{}:{}".format(get_column_letter(col)+str(row),get_column_letter(final_c)+str(final_r))
#    sheet.conditional_formatting.add(form_range,
#                                     ColorScaleRule(start_type='min', start_color='4F81BD',
#                                                    mid_type='num', mid_value=1, mid_color='FFFFFF',
#                                                    end_type='percentile', end_value=80, end_color='F8696B' ))

#"""
#Write time summary matrix
#"""
#def __writeTimeSummaryMatrix(sheet, data, row, col):
#    from TestSuite import _TIME
#    test_col = col - 1
#    test_row = row - 1
#    final_r = row+len(data.keys())
#    final_c = col+len(data.keys())
#    for r in range(row,row+len(data.keys())):
#        for c in range(col,col+len(data.keys())):
#            t2 = sheet.cell(row = r, column = test_col).value
#            t1 = sheet.cell(row = test_row, column = c).value
#            if t1 == t2:
#                sheet.cell(row = r, column = c).value = "-"
#            else:
#                r2 =[__SR_FORMAT.format(sheet=parseSheetLayer(t1)[0],cell=cl) for cl in data[t1][_TIME]]
#                r1 =[__SR_FORMAT.format(sheet=parseSheetLayer(t2)[0],cell=cl) for cl in data[t2][_TIME]]
#                sheet.cell(row = r, column = c).value = __S_TIME_FORMAT.format(*(r1+r2))
#                sheet.cell(row = r, column = c).style = 'Percent'
#            sheet.cell(row=r,column=c).alignment = xl.styles.Alignment(horizontal='center')
#    # Set conditional coloring
#    form_range = "{}:{}".format(get_column_letter(col)+str(row),get_column_letter(final_c)+str(final_r))
#    sheet.conditional_formatting.add(form_range,
#                                     ColorScaleRule(start_type='min', start_color='9BDE55',#'63BE7B',
#                                                    mid_type='num', mid_value=1, mid_color='FFFFFF',
#                                                    end_type='percentile', end_value=80, end_color='00BBEF'))#'4F81BD' ))

#"""
#Write psnr summary matrix
#"""
#def __writePSNRSummaryMatrix(sheet, data, row, col):
#    from TestSuite import _PSNR
#    test_col = col - 1
#    test_row = row - 1
#    final_r = row+len(data.keys())
#    final_c = col+len(data.keys())
#    for r in range(row,row+len(data.keys())):
#        for c in range(col,col+len(data.keys())):
#            t2 = sheet.cell(row = r, column = test_col).value
#            t1 = sheet.cell(row = test_row, column = c).value
#            if t1 == t2:
#                sheet.cell(row = r, column = c).value = 0#"-"
#            else:
#                r2 =[__SR_FORMAT.format(sheet=parseSheetLayer(t1)[0],cell=cl) for cl in data[t1][_PSNR]]
#                r1 =[__SR_FORMAT.format(sheet=parseSheetLayer(t2)[0],cell=cl) for cl in data[t2][_PSNR]]
#                sheet.cell(row = r, column = c).value = __S_PSNR_FORMAT.format(*(r1+r2))
#            sheet.cell(row = r, column = c).style = 'Comma'
#            sheet.cell(row=r,column=c).alignment = xl.styles.Alignment(horizontal='center')
#    # Set conditional coloring
#    form_range = "{}:{}".format(get_column_letter(col)+str(row),get_column_letter(final_c)+str(final_r))
#    sheet.conditional_formatting.add(form_range,
#                                     ColorScaleRule(start_type='min', start_color='FF772A',
#                                                    mid_type='num', mid_value=0, mid_color='FFFFFF',
#                                                    end_type='max', end_color='9BDE55' ))


